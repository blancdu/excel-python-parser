import string
from distutils.dir_util import copy_tree
import openpyxl
import os
import re
import time

def col2num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num


# copy src folder to dst
print('dst 폴더 생성 시작')
from_file_path = './src' # 복사할 폴더
to_file_path = './dist' # 복사 위치
copy_tree(from_file_path, to_file_path)
print('dst 폴더 생성 완료')

# 통합본 파일
# if 인게임 스트링, (String Key, col, row) 정렬 필요
main_file_name = 'sorted_translated.xlsx'
main_sheet_name = '인게임 스트링' 

print('통합본 파일 로딩 시작')
main_wb = openpyxl.load_workbook(filename=main_file_name, data_only=True)
main_ws = main_wb[main_sheet_name]
main_row_idx = 2
# main_row_idx = 950
print('통합본 파일 로딩 완료')

# 변환결과 파일
result_wb = openpyxl.Workbook()
sum_ws = result_wb.active
sum_ws.title = '요약'
sum_ws.append([
    '엑셀파일명',
    '개별 파일 내 유효 셀 수',
    '통합본 내 레코드 수',
    '통합본 내 중복 의심 레코드 수',
    '통합본 레코드 위치의 유효 셀',
    '변환완료',
    '통합본 source_cn 값 불일치',
    '통합본에 누락된 셀',
    '통합본에만 있는 정크 레코드'
    ]
)
result_wb.create_sheet('값불일치')
mismatch_ws = result_wb['값불일치']
mismatch_ws.append([
    '[통합본]String Key',
    '[통합본]Description(좌표)',
    '[통합본]Source_CN(셀값)',
    '[개별파일]실제 셀 값',
    ]
)
result_wb.create_sheet('통합본누락')
under_ws = result_wb['통합본누락']
under_ws.append([
    'String Key(예시)',
    'Description',
    'Source_CN(번역 할 필요 없으면 OK)',
    ]
)
result_wb.create_sheet('통합본정크')
over_ws = result_wb['통합본정크']
over_ws.append([
    'String Key(예시)',
    'Description',
    'Source_CN',
    'Target_KR'
    ]
)

# 개별 xlsx 파일 확인
xlsx_file_list = []
dist_file_names = os.listdir('./dist/')
for v in dist_file_names:
    name, ext = os.path.splitext(v)
    if ext == '.xlsx':
        xlsx_file_list.append(name)
xlsx_file_list.sort()


pattern = re.compile("_x000D_", re.IGNORECASE)
start_time, fps = time.perf_counter(), 0
for xlsx_file_name in xlsx_file_list:
    wb = openpyxl.load_workbook(filename='./dist/'+xlsx_file_name+'.xlsx', data_only=True)
    ws = wb.worksheets[0]
    print(f'{xlsx_file_name} 로드 완료')
    valid_cells = set()
    for row in ws:
        for cell in row:
            v = cell.value
            if v:
                v = re.sub(pattern, " ", v).strip()
                if v:
                    valid_cells.add(cell.coordinate)
                    # print(cell.coordinate, '['+v+']')
    

    # '개별 파일 내 유효 셀 수',
    # '통합본 내 레코드 수',
    # '통합본 내 중복 의심 레코드 수',
    # '통합본 레코드 위치의 유효 셀',
    # '변환완료',
    # '통합본 source_cn 값 불일치',
    # '통합본에 누락된 셀',
    # '통합본에만 있는 정크 레코드'
    # valid_cells_cnt = len(valid_cells)
    # record_cnt = 0
    # duplicated_record_cnt = 0
    # coordi_match_cnt = 0
    # success_cnt = 0
    # mismatch_cnt = 0
    # under_cnt = 0
    # over_cnt = 0
    valid_cells_cnt, record_cnt, duplicated_record_cnt, coordi_match_cnt, success_cnt, mismatch_cnt, under_cnt, over_cnt  = len(valid_cells), 0, 0, 0, 0, 0, 0, 0
    

    while True:
        main_row = main_ws[main_row_idx]
        filename = re.sub(pattern, " ", main_row[0].value).strip()
        coordi = re.sub(pattern, " ", main_row[1].value).strip()
        source_cn = re.sub(pattern, " ", main_row[2].value).strip()
        target_kr = re.sub(pattern, " ", main_row[3].value).strip()
        fps += 1
        end_time = time.perf_counter() - start_time

        print(f'[{main_row_idx}] fps: {fps/end_time:.03f}\n{filename} {coordi} {source_cn} {target_kr}')
        

        if xlsx_file_name < filename:
            break
        if xlsx_file_name > filename: #정크
            print(f'[{filename}] 파일을 찾을 수 없습니다.')
            over_cnt += 1
            over_ws.append([filename, coordi, source_cn, target_kr])
            main_row_idx += 1
            continue
        
        record_cnt += 1
        target_value = re.sub(pattern, " ", ws[coordi].value).strip()
        if target_value == source_cn:
            ws[coordi] = target_kr
            success_cnt += 1
            try:
                valid_cells.remove(coordi)
            except KeyError:
                duplicated_record_cnt += 1
                print(f'[KeyError] 같은 셀을 두 번 수정함: [{coordi}]{valid_cells}')
        else:
            mismatch_cnt += 1
            mismatch_ws.append([filename, coordi, source_cn, target_value])
            print(f'[{filename}] {coordi} 위치의 string 값이 source_cn: "{source_cn}"과 일치하지 않아 변환하지 못하였습니다.')
        
        main_row_idx += 1
        

    if valid_cells:
        under_cnt += len(valid_cells)
        for co in valid_cells:
            under_ws.append([filename, co, ws[co].value])
        print(f'{xlsx_file_name}파일에서 변환되지 않은 셀이 있습니다: {valid_cells}')

    coordi_match_cnt = success_cnt + mismatch_cnt
    # 개별 파일 저장
    wb.save('./dist/'+xlsx_file_name+'.xlsx')
    # result sum 레코드 추가
    sum_ws.append([xlsx_file_name, valid_cells_cnt, record_cnt, duplicated_record_cnt, coordi_match_cnt, success_cnt, mismatch_cnt, under_cnt, over_cnt])
    result_wb.save('result.xlsx')